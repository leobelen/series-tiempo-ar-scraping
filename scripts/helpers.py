#!/usr/bin/env python
# -*- coding: utf-8 -*-

"""Funciones auxiliares"""

from __future__ import unicode_literals
from __future__ import print_function
from __future__ import with_statement
import os
from openpyxl import load_workbook
import logging
import logging.config


def ensure_dir_exists(directory):
    """Se asegura de que un directorio exista."""
    if not os.path.exists(directory):
        os.makedirs(directory)


def get_ws_case_insensitive(wb, title):
    """Devuelve una hoja en un workbook sin importar mayúsculas/minúsculas."""
    return wb[find_ws_name(wb, title)]


def find_ws_name(wb, name):
    """Busca una hoja en un workbook sin importar mayúsculas/minúsculas."""
    if type(wb) == str or type(wb) == unicode:
        wb = load_workbook(wb, read_only=True, data_only=True)

    for sheetname in wb.sheetnames:
        if sheetname.lower() == name.lower():
            return sheetname

    return None


def row_from_cell_coord(coord):
    return int(filter(lambda x: x.isdigit(), coord))


def get_logger(name=__name__):
    logger = logging.getLogger(name)
    logger.setLevel(logging.DEBUG)
    ch = logging.StreamHandler()
    ch.setLevel(logging.DEBUG)
    logging_formatter = logging.Formatter(
        '%(asctime)s - %(name)s - %(levelname)s - %(message)s')
    ch.setFormatter(logging_formatter)
    # logger.addHandler(ch)

    return logger